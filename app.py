from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Configuration
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}

# Global variables for storing processed data
processed_appointments = []
appointments_filename = ""
excel_data = {}

def allowed_excel_file(filename):
    """Check if the uploaded file has an allowed Excel extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXCEL_EXTENSIONS

def process_excel_file(file_stream):
    """Process uploaded Excel file and extract Patient ID, Remark, and Agent Name data.
    Checks all sheets to find the one with required columns."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_stream)
        ws = None
        patient_id_col = None
        remark_col = None
        agent_name_col = None
        header_row = 1
        
        # Try all sheets to find the one with Patient ID, Remark, and Agent Name columns
        for sheet_name in wb.sheetnames:
            current_ws = wb[sheet_name]
            
            # Look for headers in the first few rows
            for row_num in range(1, min(6, current_ws.max_row + 1)):
                temp_patient_id_col = None
                temp_remark_col = None
                temp_agent_name_col = None
                
                for col in range(1, current_ws.max_column + 1):
                    cell_value = str(current_ws.cell(row=row_num, column=col).value or '').strip().lower()
                    cell_value_clean = cell_value.replace(' ', '').replace('_', '').replace('-', '').replace('.', '')
                    
                    if ('patient' in cell_value_clean and 'id' in cell_value_clean) or cell_value_clean == 'pid':
                        temp_patient_id_col = col
                    elif 'remark' in cell_value_clean:
                        temp_remark_col = col
                    elif 'agent' in cell_value_clean and 'name' in cell_value_clean:
                        temp_agent_name_col = col
                
                # If we found Patient ID and Remark (Agent Name is optional), use this sheet
                if temp_patient_id_col and temp_remark_col:
                    patient_id_col = temp_patient_id_col
                    remark_col = temp_remark_col
                    agent_name_col = temp_agent_name_col
                    header_row = row_num
                    ws = current_ws
                    break
            
            if ws:
                break
        
        # If still not found, try active sheet
        if not ws:
            ws = wb.active
            for col in range(1, ws.max_column + 1):
                cell_value = str(ws.cell(row=1, column=col).value or '').strip().lower()
                cell_value_clean = cell_value.replace(' ', '').replace('_', '').replace('-', '').replace('.', '')
                
                if ('patient' in cell_value_clean and 'id' in cell_value_clean) or cell_value_clean == 'pid':
                    patient_id_col = col
                elif 'remark' in cell_value_clean:
                    remark_col = col
                elif 'agent' in cell_value_clean and 'name' in cell_value_clean:
                    agent_name_col = col
        
        if not patient_id_col:
            sheet_names = ', '.join(wb.sheetnames)
            raise Exception(f"Patient ID column not found in Excel file. Checked sheets: {sheet_names}")
        
        if not remark_col:
            sheet_names = ', '.join(wb.sheetnames)
            raise Exception(f"Remark column not found in Excel file. Checked sheets: {sheet_names}")
        
        # Extract data - now returns list of records for each patient ID
        excel_data = {}
        data_start_row = header_row + 1
        for row in range(data_start_row, ws.max_row + 1):  # Skip header row
            patient_id = str(ws.cell(row=row, column=patient_id_col).value or '').strip()
            remark = str(ws.cell(row=row, column=remark_col).value or '').strip()
            agent_name = str(ws.cell(row=row, column=agent_name_col).value or '').strip() if agent_name_col else ''
            
            # Clean up Patient ID - remove .0 if it's a float
            if patient_id.endswith('.0'):
                patient_id = patient_id[:-2]
            
            if patient_id:  # Only add non-empty patient IDs
                if patient_id not in excel_data:
                    excel_data[patient_id] = []
                
                excel_data[patient_id].append({
                    'remark': remark,
                    'agent_name': agent_name
                })
        
        return excel_data
        
    except Exception as e:
        raise Exception(f"Error processing Excel file: {str(e)}")

def process_appointments_excel(file_stream):
    """Read an appointments Excel and return list of appointment dicts with all columns.
    
    Only requires 'Pat ID' column. All other columns are preserved as-is.
    Checks all sheets to find the one with Pat ID column.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_stream)
    ws = None
    headers = []
    pat_id_col = None
    header_row = 1
    
    # Try all sheets to find the one with Pat ID column
    for sheet_name in wb.sheetnames:
        current_ws = wb[sheet_name]
        
        # Try to find header row (check first 5 rows)
        for row_num in range(1, min(6, current_ws.max_row + 1)):
            temp_headers = []
            for col in range(1, current_ws.max_column + 1):
                raw = current_ws.cell(row=row_num, column=col).value
                name = (str(raw or '')).strip()
                temp_headers.append(name)
            
            # Check if this row looks like headers (has a Pat ID column)
            for i, header in enumerate(temp_headers):
                header_lower = header.lower().replace(' ', '').replace('_', '').replace('-', '').replace('.', '')
                # Check for various patterns: pat id, patient id, patientid, patid, etc.
                if ('pat' in header_lower and 'id' in header_lower) or header_lower == 'pid':
                    headers = temp_headers
                    header_row = row_num
                    pat_id_col = i + 1  # 1-based column index
                    ws = current_ws
                    break
            
            if pat_id_col:
                break
        
        if pat_id_col:
            break
    
    # If still not found, use active sheet and check again
    if pat_id_col is None:
        ws = wb.active
        headers = []
        for col in range(1, ws.max_column + 1):
            raw = ws.cell(row=1, column=col).value
            name = (str(raw or '')).strip()
            headers.append(name)
        
        for i, header in enumerate(headers):
            header_lower = header.lower().replace(' ', '').replace('_', '').replace('-', '').replace('.', '')
            if ('pat' in header_lower and 'id' in header_lower) or header_lower == 'pid':
                pat_id_col = i + 1
                header_row = 1
                break
    
    if pat_id_col is None:
        # Provide helpful error message with found columns
        sheet_names = ', '.join(wb.sheetnames)
        found_columns = ', '.join([f"'{h}'" for h in headers if h]) or 'none'
        raise Exception(f"Pat ID column not found in appointments Excel. Checked sheets: {sheet_names}. Found columns: {found_columns}. Please ensure there's a column containing 'Pat ID', 'Patient ID', or similar.")

    # Read all rows starting after the header row
    appointments = []
    data_start_row = header_row + 1
    for row in range(data_start_row, ws.max_row + 1):
        record = {}
        
        # Read all columns
        for col, header in enumerate(headers, 1):
            value = ws.cell(row=row, column=col).value
            record[header] = '' if value is None else str(value)
        
        # Normalize Patient ID to string without trailing .0
        pat_id_value = record.get(headers[pat_id_col - 1], '')
        pid = str(pat_id_value).strip()
        if pid.endswith('.0'):
            pid = pid[:-2]
        record['Pat ID'] = pid  # Standardize the key name
        
        # Ensure Remark and Agent Name exist
        if 'Remark' not in record:
            record['Remark'] = ''
        if 'Agent Name' not in record:
            record['Agent Name'] = ''
        
        # Skip empty rows (no Pat ID)
        if pid:
            appointments.append(record)

    return appointments

def update_appointments_with_remarks(appointments, excel_data):
    """Update appointments with remarks and agent names from Excel data based on Patient ID matching.
    Creates separate rows for each match when Patient ID appears multiple times."""
    updated_appointments = []
    updated_count = 0
    
    for appointment in appointments:
        patient_id = str(appointment.get('Pat ID', '')).strip()
        matches_found = False
        
        # Try exact match first
        if patient_id and patient_id in excel_data:
            # Create a separate row for each match
            for match_data in excel_data[patient_id]:
                new_appointment = appointment.copy()  # Copy all original data
                new_appointment['Remark'] = match_data['remark']
                new_appointment['Agent Name'] = match_data['agent_name']
                updated_appointments.append(new_appointment)
                updated_count += 1
                matches_found = True
        # Try with .0 suffix (in case Excel has float format)
        elif patient_id and f"{patient_id}.0" in excel_data:
            for match_data in excel_data[f"{patient_id}.0"]:
                new_appointment = appointment.copy()  # Copy all original data
                new_appointment['Remark'] = match_data['remark']
                new_appointment['Agent Name'] = match_data['agent_name']
                updated_appointments.append(new_appointment)
                updated_count += 1
                matches_found = True
        
        # If no matches found, add original appointment with empty remark and agent name
        if not matches_found:
            appointment['Remark'] = ''
            appointment['Agent Name'] = ''
            updated_appointments.append(appointment)
    
    return updated_appointments, updated_count

def create_excel_from_appointments(appointments, filename):
    """Create Excel file from processed appointment data with all columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Appointment Data"
    
    if not appointments:
        return wb
    
    # Get all unique headers from all appointments
    all_headers = set()
    for appointment in appointments:
        all_headers.update(appointment.keys())
    
    # Convert to list and ensure Pat ID, Remark, and Agent Name are at the end for visibility
    headers = list(all_headers)
    if 'Pat ID' in headers:
        headers.remove('Pat ID')
    if 'Remark' in headers:
        headers.remove('Remark')
    if 'Agent Name' in headers:
        headers.remove('Agent Name')
    headers.extend(['Pat ID', 'Remark', 'Agent Name'])  # Put these at the end
    
    # Set headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add appointment data
    for row, appointment in enumerate(appointments, 2):
        for col, header in enumerate(headers, 1):
            value = appointment.get(header, '')
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

@app.route('/')
def index():
    """Main page with file upload form."""
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file uploads - two Excel inputs: appointments and remarks."""
    global processed_appointments, appointments_filename, excel_data

    appointments_file = request.files.get('appointments_file')
    remarks_file = request.files.get('remarks_file')

    # Require both Excel files
    if not appointments_file or appointments_file.filename == '':
        flash('Please upload the Appointments Excel file.')
        return redirect(url_for('index'))

    if not remarks_file or remarks_file.filename == '':
        flash('Please upload the Remarks Excel file.')
        return redirect(url_for('index'))

    if not allowed_excel_file(appointments_file.filename):
        flash('Invalid Appointments Excel file type. Please upload .xlsx or .xls')
        return redirect(url_for('index'))

    if not allowed_excel_file(remarks_file.filename):
        flash('Invalid Remarks Excel file type. Please upload .xlsx or .xls')
        return redirect(url_for('index'))

    try:
        # Process appointments Excel directly from memory
        appointments_filename_raw = secure_filename(appointments_file.filename)
        processed_appointments = process_appointments_excel(appointments_file)
        appointments_filename = os.path.splitext(appointments_filename_raw)[0]

        flash(f'Successfully processed appointments with {len(processed_appointments)} rows.')
    except Exception as e:
        flash(f'Error processing Appointments Excel: {str(e)}')
        return redirect(url_for('index'))

    # Process remarks Excel
    try:
        excel_data = process_excel_file(remarks_file)
        updated_appointments, updated_count = update_appointments_with_remarks(processed_appointments, excel_data)
        
        # Update the global processed_appointments with the new data
        processed_appointments = updated_appointments

        flash(f'Successfully updated {updated_count} appointments with remarks and agent names. Total rows: {len(processed_appointments)}')
    except Exception as e:
        flash(f'Error processing Remarks Excel: {str(e)}')
        return redirect(url_for('index'))

    return redirect(url_for('results'))


@app.route('/results')
def results():
    """Show extracted appointment data."""
    global processed_appointments, appointments_filename
    
    if not processed_appointments:
        flash('No data found. Please upload Excel files first.')
        return redirect(url_for('index'))
    
    return render_template('results.html', 
                         appointments=processed_appointments, 
                         filename=appointments_filename,
                         total_appointments=len(processed_appointments))

@app.route('/download')
def download_excel():
    """Download the Excel file."""
    global processed_appointments, appointments_filename
    
    if not processed_appointments:
        flash('No data to download. Please upload Excel files first.')
        return redirect(url_for('index'))
    
    try:
        # Create Excel file
        excel_buffer = create_excel_from_appointments(processed_appointments, appointments_filename)
        
        return send_file(excel_buffer, 
                        as_attachment=True, 
                        download_name=f'{appointments_filename}_appointments.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        flash(f'Error creating Excel file: {str(e)}')
        return redirect(url_for('index'))

if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    port = int(os.environ.get('PORT', 8080))
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
